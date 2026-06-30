#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
XLSX 格式化模块
===============
将表格格式化为适合 A4 打印的样式：自动换行、列宽适配、横纵判断、打印缩放。

用法:
  from scripts.xlsx_formatter import format_xlsx
  output = format_xlsx("input.xlsx")  # 返回格式化后的 xlsx 路径
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# ── 常量 ────────────────────────────────────────────
A4_PORTRAIT_MAX_COLS = 80   # 竖版最多容纳的总列宽（字符单位）
MAX_COL_WIDTH = 50
MIN_COL_WIDTH = 4

HEADER_FONT = Font(name="微软雅黑", size=10, bold=True)
BODY_FONT   = Font(name="微软雅黑", size=9)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
BODY_ALIGN   = Alignment(horizontal="left", vertical="center", wrap_text=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ── 工具函数 ────────────────────────────────────────

def char_width(text):
    """估算字符串显示宽度：CJK=2, ASCII=1"""
    if not text:
        return 0
    w = 0
    for ch in str(text):
        if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
            w += 2
        else:
            w += 1
    return w


def _calc_col_width(ws, col_idx, sample_rows=200):
    """计算第 col_idx 列的最优宽度"""
    max_w = MIN_COL_WIDTH
    max_row = min(_real_max_row(ws), sample_rows)
    for row in range(1, max_row + 1):
        val = ws.cell(row, col_idx).value
        if val is not None:
            w = char_width(str(val))
            if w > max_w:
                max_w = w
    return min(max_w + 2, MAX_COL_WIDTH)


def _real_max_row(ws):
    """反向扫描检测真实数据行数（openpyxl 的 max_row 可能包含格式化空行）"""
    max_col = ws.max_column or 0
    for row in range(ws.max_row or 0, 0, -1):
        if any(ws.cell(row, c).value is not None for c in range(1, max_col + 1)):
            return row
    return 0


def _format_one_sheet(ws):
    """格式化单个 sheet"""
    max_row = _real_max_row(ws)
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return

    # 样式
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            cell.border = THIN_BORDER
            if row == 1:
                cell.font = HEADER_FONT
                cell.alignment = HEADER_ALIGN
            else:
                cell.font = BODY_FONT
                cell.alignment = BODY_ALIGN

    # 列宽
    total_width = 0
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        w = _calc_col_width(ws, col)
        ws.column_dimensions[col_letter].width = w
        total_width += w

    # 打印设置
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ps = ws.page_setup
    ps.fitToWidth = 1
    ps.fitToHeight = 0
    ps.paperSize = 9  # A4
    ps.orientation = "landscape" if total_width > A4_PORTRAIT_MAX_COLS else "portrait"

    ws.page_margins = openpyxl.worksheet.page.PageMargins(
        left=0.5, right=0.5, top=0.6, bottom=0.6,
        header=0.3, footer=0.3,
    )

    return {"rows": max_row, "cols": max_col, "total_width": total_width,
            "orientation": ps.orientation}


def format_xlsx(input_path, output_path=None, sheets=None):
    """
    格式化 xlsx 文件。
    
    Args:
        input_path: 源 xlsx 路径
        output_path: 输出路径（默认同目录 _格式化 后缀）
        sheets: 要处理的 sheet 名列表（默认全部）
    
    Returns:
        dict: {"path": output_path, "sheets": [{name, rows, cols, orientation}, ...]}
    """
    input_path = Path(input_path)
    if output_path is None:
        output_path = input_path.parent / (input_path.stem + "_格式化.xlsx")
    else:
        output_path = Path(output_path)

    wb = openpyxl.load_workbook(input_path)
    all_sheets = wb.sheetnames
    target = sheets or all_sheets

    results = []
    for sn in target:
        if sn not in all_sheets:
            continue
        ws = wb[sn]
        info = _format_one_sheet(ws)
        if info:
            info["name"] = sn
            results.append(info)

    wb.save(str(output_path))
    wb.close()

    return {"path": str(output_path), "sheets": results}


def analyze_xlsx(input_path):
    """
    快速分析 xlsx 结构（不修改文件）。
    
    Returns:
        dict: {"sheets": [{name, rows, cols}], "total_sheets": N}
    """
    input_path = Path(input_path)
    wb = openpyxl.load_workbook(input_path, data_only=True)

    sheets_info = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        max_col = ws.max_column or 0
        real_rows = _real_max_row(ws)
        sheets_info.append({"name": sn, "rows": real_rows, "cols": max_col})

    wb.close()
    return {"sheets": sheets_info, "total_sheets": len(sheets_info)}
